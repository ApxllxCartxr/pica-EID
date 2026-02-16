"""Excel (.xlsx) export/import service."""

import os
from datetime import datetime
from typing import BinaryIO, Tuple
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session, joinedload

from app.models.user import User, UserStatus, UserCategory
from app.models.role import UserRole
from app.config import settings


class ExcelService:
    """Service for exporting/importing user data to/from Excel files."""

    HEADERS = ["ULID", "Name", "Email", "Category", "Status", "Roles", "Domain", "Division", "Start Date", "End Date"]

    def __init__(self, db: Session):
        self.db = db

    def export_users(self) -> Tuple[str, int]:
        """
        Export data to an Excel file with multiple tabs.
        Returns (filename, total_record_count).
        """
        os.makedirs(settings.EXCEL_EXPORT_DIR, exist_ok=True)

        wb = Workbook()
        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        total_count = 0

        # --- Helper to create and populate a sheet ---
        def create_sheet(title: str, query_filter, include_deleted=False):
            ws = wb.create_sheet(title=title)
            ws.append(self.HEADERS)
            
            # Style headers
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Base query
            query = self.db.query(User).options(
                joinedload(User.domain),
                joinedload(User.division),
                joinedload(User.user_roles).joinedload(UserRole.role),
                joinedload(User.internship),
            )

            # Apply filters
            if include_deleted:
                query = query.filter(User.deleted_at != None)
            else:
                query = query.filter(User.deleted_at == None)
                if query_filter is not None:
                    query = query.filter(query_filter)
            
            users = query.order_by(User.created_at).all()
            
            for user in users:
                roles = ", ".join(
                    ur.role.name for ur in user.user_roles
                    if ur.removed_at is None and ur.role.is_active
                )
                ws.append([
                    user.ulid,
                    user.name,
                    user.email,
                    user.category.value,
                    user.status.value,
                    roles,
                    user.domain.name if user.domain else "",
                    user.division.name if user.division else "",
                    str(user.internship.start_date) if user.internship else "",
                    str(user.internship.end_date) if user.internship else "",
                ])

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)
            
            return len(users)

        # 1. Active Employees
        total_count += create_sheet(
            "Active Employees", 
            (User.category == UserCategory.EMPLOYEE) & (User.status == UserStatus.ACTIVE)
        )

        # 2. Interns
        total_count += create_sheet(
            "Interns", 
            (User.category == UserCategory.INTERN) & (User.status == UserStatus.ACTIVE)
        )

        # 3. Inactive Employees/Interns
        total_count += create_sheet(
            "Inactive Users", 
            User.status.in_([UserStatus.INACTIVE, UserStatus.EXPIRED])
        )

        # 4. Deleted Users
        total_count += create_sheet(
            "Deleted Users", 
            None, 
            include_deleted=True
        )

        # 5. Domains
        from app.models.domain import Domain
        ws_domains = wb.create_sheet(title="Domains")
        ws_domains.append(["ID", "Name", "Description", "Active User Count"])
        domains = self.db.query(Domain).all()
        for d in domains:
             count = self.db.query(User).filter(User.domain_id == d.id, User.deleted_at == None).count()
             ws_domains.append([d.id, d.name, d.description, count])
        
        # 6. Departments (Divisions)
        from app.models.division import Division
        ws_divs = wb.create_sheet(title="Departments")
        ws_divs.append(["ID", "Name", "Description", "Active User Count"])
        divisions = self.db.query(Division).all()
        for d in divisions:
             count = self.db.query(User).filter(User.division_id == d.id, User.deleted_at == None).count()
             ws_divs.append([d.id, d.name, d.description, count])

        # 7. Users by Domain
        ws_by_domain = wb.create_sheet(title="Users by Domain")
        ws_by_domain.append(self.HEADERS)
        users_by_domain = self.db.query(User).options(
                joinedload(User.domain),
                joinedload(User.division),
                joinedload(User.user_roles).joinedload(UserRole.role),
                joinedload(User.internship),
            ).filter(User.deleted_at == None).order_by(User.domain_id, User.name).all()
        
        for user in users_by_domain:
             roles = ", ".join(ur.role.name for ur in user.user_roles if ur.removed_at is None)
             ws_by_domain.append([
                user.ulid, user.name, user.email, user.category.value, user.status.value, roles,
                user.domain.name if user.domain else "Unassigned",
                user.division.name if user.division else "",
                str(user.internship.start_date) if user.internship else "",
                str(user.internship.end_date) if user.internship else "",
            ])

        # 8. Users by Department
        ws_by_div = wb.create_sheet(title="Users by Department")
        ws_by_div.append(self.HEADERS)
        users_by_div = self.db.query(User).options(
                joinedload(User.domain),
                joinedload(User.division),
                joinedload(User.user_roles).joinedload(UserRole.role),
                joinedload(User.internship),
            ).filter(User.deleted_at == None).order_by(User.division_id, User.name).all()
        
        for user in users_by_div:
             roles = ", ".join(ur.role.name for ur in user.user_roles if ur.removed_at is None)
             ws_by_div.append([
                user.ulid, user.name, user.email, user.category.value, user.status.value, roles,
                user.domain.name if user.domain else "",
                user.division.name if user.division else "Unassigned",
                str(user.internship.start_date) if user.internship else "",
                str(user.internship.end_date) if user.internship else "",
            ])

        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"prismid_export_{timestamp}.xlsx"
        filepath = os.path.join(settings.EXCEL_EXPORT_DIR, filename)
        wb.save(filepath)

        return filename, total_count

    def import_users(self, file: BinaryIO, admin_id: int) -> dict:
        """
        Import user updates from an uploaded Excel file.
        Only updates existing users (matched by User ID).
        Returns dict with counts and errors.
        """
        wb = load_workbook(file, read_only=True)
        ws = wb.active

        result = {"updated": 0, "skipped": 0, "errors": []}

        # Find header row
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and row[0] == "ULID":
                header_row = row_idx
                break

        if header_row is None:
            raise ValueError("Could not find header row with 'ULID' column")

        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not row or not row[0]:
                continue

            ulid_value = str(row[0]).strip()
            user = self.db.query(User).filter(User.ulid == ulid_value).first()

            if not user:
                result["errors"].append(f"Row {row_idx}: ULID '{ulid_value}' not found")
                result["skipped"] += 1
                continue

            try:
                # Update name if provided and different
                if len(row) > 1 and row[1] and str(row[1]).strip() != user.name:
                    user.name = str(row[1]).strip()

                # Update email if provided and different
                if len(row) > 2 and row[2] and str(row[2]).strip() != user.email:
                    user.email = str(row[2]).strip()

                # Update status if valid
                if len(row) > 4 and row[4]:
                    new_status = str(row[4]).strip().upper()
                    if new_status in [s.value for s in UserStatus]:
                        user.status = UserStatus(new_status)

                user.updated_at = datetime.utcnow()
                result["updated"] += 1

            except Exception as e:
                result["errors"].append(f"Row {row_idx}: {str(e)}")
                result["skipped"] += 1

        self.db.commit()
        wb.close()
        return result
